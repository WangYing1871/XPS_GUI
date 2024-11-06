import sys
import os
import matplotlib
import json
import pandas as pd
import openpyxl
from openpyxl import load_workbook
# import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvas  # pyqt5的画布
from matplotlib.font_manager import FontProperties
import datetime
# from PyQt5 import QtWidgets
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QGraphicsScene, QVBoxLayout, QWidget, QPushButton, QLineEdit, QLabel, QDialog, QDialogButtonBox, QFormLayout
from Ui_DataTransmission import Ui_DataTransmisson
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QTimer,QEventLoop
from PyQt5.QtNetwork import QUdpSocket, QHostAddress

matplotlib.use("Qt5Agg")  # 声明使用pyqt5
matplotlib.rcParams['font.family'] = 'SimHei'  # 'SimHei' 是一种常用的中文黑体
matplotlib.rcParams['font.size'] = 10
matplotlib.rcParams['axes.unicode_minus'] = False  # 正确显示负号


class MyMatplotlibFigure(FigureCanvas):  # 创建一个绘图类,主要用于绘制能谱图
    def __init__(self, width=7.1, height=5.51, dpi=1200):
        fig = Figure(figsize=(width, height), dpi=dpi)
        self.axes = fig.add_subplot(111)
        fig.subplots_adjust(left=0.07, right=0.99, top=1, bottom=0.05)
        self.font_prop = FontProperties(family='SimHei', size=8)
        self.axes_channels = 32
        super(MyMatplotlibFigure, self).__init__(fig)

    def draw_bar_chart(self, data):
        self.axes.clear()  # 清除之前的绘图

        channel_names = list(data.keys())
        channel_values = list(data.values())
        bars = self.axes.bar(channel_names, channel_values, color=(29/255, 100/255, 87/255), edgecolor='black')

        # 显示总计数
        total_counts = sum(channel_values)
        self.axes.text(0.98, 0.98, f'总计数: {total_counts}', ha='right', va='top', transform=self.figure.transFigure, fontsize=12, fontproperties=self.font_prop, bbox=dict(facecolor='white', alpha=0.8))

        # 设置x轴标签
        self.axes.set_xticks(range(len(channel_names)))  # 设置x轴刻度位置
        self.axes.set_xticklabels(range(self.axes_channels), fontproperties=self.font_prop)  # 设置x轴标签

        # 在x轴刻度下方添加计数标签
        for i, value in enumerate(channel_values):
            self.axes.text(i, -0.05 * max(channel_values), f'{value}', ha='center', va='top', fontsize=6, fontproperties=self.font_prop)

        # 设置y轴标签的字体属性
        for label in self.axes.get_yticklabels():
            label.set_fontproperties(self.font_prop)

        y_max = max(channel_values) if max(channel_values) > 0 else 1
        self.axes.set_ylim(-0.1 * y_max, y_max * 1.1)  # 设置y轴范围

        self.draw()

        width, height = self.get_width_height()
        buf = self.buffer_rgba()
        qimage = QImage(buf, width, height, QImage.Format_RGBA8888)
        pixmap = QPixmap.fromImage(qimage)

        return pixmap

    def clear_plot(self):
        self.axes.clear()
        # self.draw()

class ReceiverThread(QThread):  # 定义一个接收数据的线程类
    dataReceived = pyqtSignal(bytes, str, int)  # 定义一个信号，用于发送接收到的数据
    ReceivedError = pyqtSignal(str)  # 定义一个信号，用于发送错误信息

    def __init__(self, socket):  # 初始化函数
        super().__init__()
        self.socket = socket  # 传入一个socket对象

    def run(self):  # 线程运行函数
        self.socket.readyRead.connect(self.readData)  # 连接信号readyRead到槽函数readData，当缓冲区有数据时，触发readyRead信号

    def readData(self):  # 读取数据
        while self.socket.hasPendingDatagrams():  # 当缓冲区有数据时
            datagram, host, port = self.socket.readDatagram(self.socket.pendingDatagramSize())  # 读取数据
            try:
                self.dataReceived.emit(datagram, host.toString(), port)  # 发送接收到的数据，发送信号，等待主线程处理，保证线程安全通信
            except Exception as e:
                self.ReceivedError.emit('开发板未在监听，请检查开发板的UDP连接: %s' % e)  # 发送错误信息，发送信号，等待主线程处理，保证线程安全通信

class SampStepTotalLengthDialog(QDialog):   # 创建一个图窗用来输入采集步长和总长度
    def __init__(self):
        super().__init__()
        
        self.setWindowTitle('输入采集步长和总长度')
        
        self.layout = QFormLayout(self)
        
        self.stepInput = QLineEdit(self)
        self.lengthInput = QLineEdit(self)
        
        self.layout.addRow('采集步长:', self.stepInput)
        self.layout.addRow('采集总长度:', self.lengthInput)
        
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        
        self.layout.addWidget(self.buttonBox)
    
    def getInputs(self):
        return float(self.stepInput.text()), float(self.lengthInput.text())

class SCurveHandler:    # 用于一键生成S曲线
    def __init__(self, Data_transmission):
        self.Data_transmission = Data_transmission
        self.current_dac_value = 0
        self.initial_dac_value = 0
        self.stop_flag = False  # 初始化标志位
                
    def measure_s_curve(self, sampling_step, total_length):
        self.stop_flag = False  # 每次开始测量前重置标志位
        initial_dac_value_str = self.Data_transmission.ThresholdLineEdit.text()  # 获取阈值初值
        self.initial_dac_value = int(initial_dac_value_str)  # 将阈值初值转化为整数
        
        AcquireTime = self.Data_transmission.AcquireTimeLineEdit.text()  # 获取采集时长
        AcquireTimeValue = int(AcquireTime)  # 将阈值转化为整数
        Delay = int(AcquireTimeValue/10 +2000) # 1倍采集时长+2s，定时器的单位是1ms
        
        for step in range(0, int(total_length), int(sampling_step)):
            if self.stop_flag:  # 检查标志位，如果被设置为停止则退出流程
                self.Data_transmission.SpectroscopyTextBrowser.append("S曲线生成已停止")
                self.Data_transmission.CommunicationTextBrowser.append("S曲线生成已停止")
                break
            
            self.current_dac_value = self.initial_dac_value + step    # 当前阈值
            
            self.Data_transmission.SpectroscopyTextBrowser.append(f"当前DAC阈值： {self.current_dac_value}")
            self.Data_transmission.CommunicationTextBrowser.append(f"当前DAC阈值： {self.current_dac_value}")

            self.retry_loop(lambda: self.Data_transmission.ThresholdConfig(self.current_dac_value), 
                    "阈值配置成功", 
                    "阈值配置信号未接收到应答包，重新发送")
            
            self.retry_loop(self.Data_transmission.CountsRest, 
                    "清零成功", 
                    "清零信号未接收到应答包，重新发送")
            
            self.Data_transmission.ifSynCtrlTriggerSuccess = 0
            self.Data_transmission.SynCtrlTrigger()
            
            loop = QEventLoop()
            QTimer.singleShot(Delay, loop.quit)
            loop.exec_()
            
            if self.Data_transmission.ifSynCtrlTriggerSuccess == 2:
                self.retry_loop(self.Data_transmission.AcquireData, 
                    "读数成功", 
                    "读数信号未接收到应答包，重新发送")
            elif self.Data_transmission.ifSynCtrlTriggerSuccess == 1:
                self.Data_transmission.CommunicationTextBrowser.append("探测器能谱采集工作异常，建议检查")
                self.Data_transmission.SpectroscopyTextBrowser.append("探测器能谱采集工作异常，建议检查")
                break
            elif self.Data_transmission.ifSynCtrlTriggerSuccess == 0:
                self.Data_transmission.SynCtrlTrigger()
                QTimer.singleShot(Delay, loop.quit)  # 设置1s的定时器，触发时退出循环
                loop.exec_()    # 开始循环并等待直到定时器触发退出循环
                if self.Data_transmission.ifSynCtrlTriggerSuccess == 0:
                    self.Data_transmission.CommunicationTextBrowser.append("中间板未能收到正确应答信号，建议检查")
                    self.Data_transmission.SpectroscopyTextBrowser.append("中间板未能收到正确应答信号，建议检查")
                    break
                elif self.Data_transmission.ifSynCtrlTriggerSuccess == 2:
                    self.retry_loop(self.Data_transmission.AcquireData, 
                    "读数成功", 
                    "读数信号未接收到应答包，重新发送")
            
            self.FileSaveToExcel()  # 保存到excel文件中
                    
    def FileSaveToExcel(self):  # 文件保存
        FilePath = self.Data_transmission.FilePathLineEdit.text()
        FileName = self.Data_transmission.FileNameLineEdit.text()

        if not FileName.endswith('.xlsx'):
            FileName += '.xlsx'

        full_path = os.path.join(FilePath, FileName)

        # 将数据转换为 DataFrame
        df = pd.DataFrame(self.Data_transmission.spectrum_data_list).T  # 转置使每一步的数据为一列
     
        try:
            # 检查并创建文件所在的目录
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
        except Exception as e:
            self.Data_transmission.SpectroscopyTextBrowser.append("文件路径错误,无法创建：%s" % e)
            self.Data_transmission.CommunicationTextBrowser.append("文件路径错误,无法创建：%s" % e)
            return
        
        # 打开或创建一个新的 Excel 文件
        new_file = False
        try:
            workbook = load_workbook(full_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            new_file = True

        # 找到下一个空列
        if new_file:
            next_col = 1
        else:
            next_col = sheet.max_column + 1

        # 写入 DAC 值到第一行
        sheet.cell(row=1, column=next_col, value=self.current_dac_value)

        # 写入通道数据到后续行
        for i, value in enumerate(df.iloc[:, -1]):
            sheet.cell(row=i + 2, column=next_col, value=value)

        # 保存文件
        workbook.save(full_path)

        self.Data_transmission.SpectroscopyTextBrowser.append(f"能谱数据已保存到 {full_path}")
        self.Data_transmission.CommunicationTextBrowser.append(f"能谱数据已保存到 {full_path}")           
            
    def retry_loop(self, action, success_message, fail_message):
        success = False
        while not success:
            action()
            if self.Data_transmission.Loop() == 1:
                success = True
                self.Data_transmission.SpectroscopyTextBrowser.append(success_message)
                self.Data_transmission.CommunicationTextBrowser.append(success_message)
            else:
                self.Data_transmission.SpectroscopyTextBrowser.append(fail_message)
                self.Data_transmission.CommunicationTextBrowser.append(fail_message)                   
        
    def stop_s_curve(self):
        self.stop_flag = True  # 设置标志位为停止    
        
class DataTransmission(QMainWindow, Ui_DataTransmisson):
    def __init__(self, parent=None):
        super(DataTransmission, self).__init__(parent)
        self.setupUi(self)  # 设置UI界面

        self.plotCanvas = MyMatplotlibFigure(width=7.1, height=5.51, dpi=1200)  # 创建一个绘图类
        self.initUI()  # 初始化能谱图窗口
        self.spectrum_data_list = []
        self.SCurveHandler = SCurveHandler(self)
        
        self.InsSN = 1  # 初始化指令帧序号为1
        self.InstrSN = format(self.InsSN, '04X')
        self.InstrReserve = 0X00  # 初始化保留字段为0
        self.str_InstrReserve = f"{self.InstrReserve:02X}"
        self.InstrPacketStart = 0XEB90  # 初始化包头为0xEB90
        self.str_InstrPacketStart = f"{self.InstrPacketStart:04X}"
        self.CtrlRegTestChannel = 0B00  # 初始化控制寄存器15-14选择测试通道，默认00为ch1，01为ch9，10为ch17，11为ch25
        self.CtrlRegFormingTime = 0B00  # 初始化控制寄存器13-12成形时间调节，默认00为40ns，01为80ns，10为120ns，11为160ns
        self.CtrlRegSerialPortRate = 0B00_00   # 初始化控制寄存器11-8串口速率模式，0000为Debug模式，0001为正常模式，0010\0011保留
        self.CtrlRegTestSignalOutputEnable = 0B00_00 # 初始化控制寄存器保留字段7-5为0, 最后一个0用于设置测试信号使能，0为禁止，1为使能
        self.CtrlRegDataMode = 0B00  # 初始化控制寄存器3-2数据模式，00为短包，01为长包,10\11保留
        self.CtrlRegWorkingMode = 0B00  # 初始化控制寄存器1-0工作模式与触发接收使能，高位控制工作模式，0为正常取数，1为电子学刻度。低位控制触发接收使能，0为禁止，1为使能
        self.ifdatareceived = 0 # 判断是否接收到指令
        dac_value = 0   # EXTENSION
        self.ifSynCtrlTriggerSuccess = 0
        
        # 信号槽连接,实现具体功能
        self.FilePathchooseButton.clicked.connect(self.openFolderDialog)  # 打开文件路径
        self.UDPBindButton.clicked.connect(self.UDPBind)  # 绑定UDP
        self.SendInsButton.clicked.connect(self.SendIns)  # 发送指令，只需要输入命令码和命令参数，也就是六位十六进制数
        self.AcquireModeConfigButton.clicked.connect(self.on_AcquireMode_changed)  # 测试模式检测
        self.InstrLoopTest.clicked.connect(self.on_InsLoopTest)  # 指令回环测试
        self.AcquireButton.clicked.connect(self.AcquireData)  # 采集数据
        self.CountsRestButton.clicked.connect(self.CountsRest)  # 计数复位
        self.ThresholdConfigButton.clicked.connect(self.ThresholdConfig)  # 阈值配置
        # self.ThresholdReadButton.clicked.connect(self.ThresholdRead)  # 阈值读取
        self.AcquireConfigButton.clicked.connect(self.AcquireConfig)  # 采集配置
        # self.AcquireReadButton.clicked.connect(self.AcquireRead)  # 采集读取
        self.CtrlRegReadButton.clicked.connect(self.CtrlRegRead)  # 控制寄存器读取
        self.FPGAResetButton.clicked.connect(self.FPGAReset)  # FPGA复位
        self.FileSaveButton.clicked.connect(self.FileSave)  # 文件保存
        self.FileDeleteButton.clicked.connect(self.FileDelete)  # 文件删除
        self.CtrlRegConfigButton.clicked.connect(self.CtrlRegConfig)    # 配置控制寄存器，包括触发接收使能，测试信号使能，工作模式选择，修改测试通道和成形时间
        self.SynCtrlTriggerBotton.clicked.connect(self.SynCtrlTrigger)  # 同步控制触发
        self.ScaleThresholdButton.clicked.connect(self.ScaleThreshold)  # 配置刻度阈值，配置刻度DAC
        self.ScaleTimeIntervalButton.clicked.connect(self.ScaleTimeInterval)    # 配置刻度时间间隔
        self.SaveConfigButton.clicked.connect(self.SaveConfig)  # 保存配置至当前文件夹
        self.LoadConfigButton.clicked.connect(self.LoadConfig)  # 导入配置文件，自动改变参数
        self.ClearSpecButton.clicked.connect(self.ClearSpec)    # 清除能谱图
        self.SampStepTotalLengthButton.clicked.connect(self.SampStepTotalLength)    # 设置测量S曲线前的采样步长与总长度
        self.SCurveButton.clicked.connect(self.SCurve)  # 一键测量S曲线
        self.StopSCurveButton.clicked.connect(self.SCurveHandler.stop_s_curve)  # 停止测量S曲线
        self.PeriodButton.clicked.connect(self.PeriodCollect)   # 周期同步触发和读数
        self.SingleChannelButton.clicked.connect(self.SingleChannelThresholdTuning)
        
    def initUI(self):  # 为能谱图初始化一个场景，建立图窗
        # 创建一个场景,初始化能谱图窗
        self.scene = QGraphicsScene(self)
        self.scene.addWidget(self.plotCanvas)
        self.SpectroscopyView.setScene(self.scene)

    def send_data(self, data):   # 发送数据
        self.InsSN += 1  # 指令帧序号加1  
        # self.CommunicationTextBrowser.append(f"指令帧: {self.InsSN}")
        # self.SpectroscopyTextBrowser.append(f"指令帧: {self.InsSN}")
        
        if self.InsSN > 65535:  # 指令帧序号超出范围，抛出异常
            self.SpectroscopyTextBrowser.append("指令溢出范围，无法发送")
            raise ValueError("指令溢出范围，无法发送")
        self.InstrSN = format(self.InsSN, '04X')  # 将指令帧序号转化为四位十六进制
        
        if isinstance(data, str):
            # 如果消息是字符串，使用 UTF-8 编码转换为 bytes
            data = data.encode('utf-8')
        elif isinstance(data, bytes):
            # 如果消息是 bytes，直接发送
            data = data
            
        try:
            self.udpSocket.writeDatagram(data, QHostAddress(self.IPAddress), self.UDPPort)  # 发送数据
        except Exception as e:
            self.CommunicationTextBrowser.append("套接字未绑定，请检查连接: %s" % e)
            self.SpectroscopyTextBrowser.append("套接字未绑定，请检查连接: %s" % e)

        self.Instruction = data.hex()  # 将指令转化为十六进制
        self.UpperInstruction = ' '.join([self.Instruction[i:i + 2].upper() for i in range(0, len(self.Instruction), 2)])  # 每两个字符之间加一个空格
        self.CommunicationTextBrowser.append("发送指令：%s" % self.UpperInstruction)
        # self.SpectroscopyTextBrowser.append("发送指令：%s" % self.UpperInstruction)
            
    def onDataReceived(self, data, host, port):  # 接收数据
        Hex_data = data.hex()
        
        self.ifdatareceived += 1    # 接收到信号则+1
        
        ls = []  # 用于存储解码后的数据
        for i in range(len(Hex_data) // 2):  # 对十六进制数据进行byte解码，且转化为大写
            hex_value = Hex_data[i * 2].upper() + Hex_data[i * 2 + 1].upper()
            ls.append(hex_value)

        OData = ' '.join(ls)
        self.CommunicationTextBrowser.append("[%s] : %s" % (self.Currenttimemessage(), OData))  # 在主界面显示解码后的数据
        self.SpectroscopyTextBrowser.append("[%s] : %s" % (self.Currenttimemessage(), OData))  # 在配置界面显示接收到的数据

        self.DataReceiveVerify(ls)  # 对接收到的数据进行校验并且解码
        
    def Currenttimemessage(self):  # 获取当前时间和消息
        # 获取当前时间
        current_time = datetime.datetime.now()
        # 格式化时间为易读的格式
        formatted_time = current_time.strftime("%H:%M:%S")
        # 组合时间和消息
        return formatted_time

    def DataReceiveVerify(self, data):  # 应答包包头与CRC校验
        ls = data
        AcquireMode = self.AcquireModeBox.currentText()
        if len(ls) == 138:
            if ls[0] == 'EB' and ls[1] == '90':  # 判断数据包头是否为0xEB90，且数据长度是否为138
                if AcquireMode == "短包":  # 判断是否为短包模式
                    # EXTENSION    
                    if all(x == '00' for x in ls[9:]):  # 检查第9位开始是否全为0，判定为短包
                        self.ShortDatadistinguish(ls)
                        
                    elif all(x == ls[9] for x in ls[10:]):    #检查是否为应答包
                        InstrCRC = ls[8] + ls[9]  # 取出数据包中的CRC码
    
                        CRCVerify = ls[2] + ls[3] + ls[4] + ls[5] + ls[6] + ls[7]  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.Insdistinguish(ls)
                        # EXTENSION
                        # if InstrCRC == CRCVerify_result:
                        #     self.Insdistinguish(ls)  # CRC校验无误时，进一步进行指令区分
                        # else:
                        #     self.CommunicationTextBrowser.append("CRC校验失败")
                        #     # raise ValueError("CRC校验失败")
                            
                    elif all(x == ls[73] for x in ls[74:]):  # 检查是否为短包，短包数据应答包长度为74  EXTENSION
                        InstrCRC = ls[72] + ls[73]

                        CRCVerify = ''.join(ls[2:72])  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.ShortDatadistinguish(ls)
                        # EXTENSION
                        # if InstrCRC == CRCVerify_result:
                        #     self.ShortDatadistinguish(ls)  # CRC校验无误时，进一步进行指令区分
                        # else:
                        #     self.CommunicationTextBrowser.append("CRC校验失败")
                        #     # raise ValueError("CRC校验失败")
                    else:
                        self.CommunicationTextBrowser.append("短包校验失败")
                
                elif AcquireMode == "长包":  # 判断是否为长包模式
                    # EXTENSION    
                    if all(x == '00' for x in ls[9:]):  # 检查第9位开始是否全为0，判定为长包
                        self.LongDatadistinguish(ls)
                        
                    elif all(x == ls[9] for x in ls[10:]):    #检查是否为应答包
                        InstrCRC = ls[8] + ls[9]  # 取出数据包中的CRC码
    
                        CRCVerify = ls[2] + ls[3] + ls[4] + ls[5] + ls[6] + ls[7]  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.Insdistinguish(ls)
                        # EXTENSION
                        # if InstrCRC == CRCVerify_result:
                        #     self.Insdistinguish(ls)  # CRC校验无误时，进一步进行指令区分
                        # else:
                        #     self.CommunicationTextBrowser.append("CRC校验失败")
                        #     # raise ValueError("CRC校验失败")
                                
                    else:   # 若不是应答包也不是短包，则是长包，长包数据应答包长度为138
                        InstrCRC = ls[136] + ls[137]

                        CRCVerify = ''.join(ls[2:136])  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.LongDatadistinguish(ls)
                        # EXTENSION
                        # if InstrCRC == CRCVerify_result:
                        #     self.LongDatadistinguish(ls)  # CRC校验无误时，进一步进行指令区分
                        #     self.CommunicationTextBrowser.append("CRC校验成功")
                        # else:
                        #     self.CommunicationTextBrowser.append("CRC校验失败")
                        #     # raise ValueError("CRC校验失败")
                elif AcquireMode == "簇团计数":  # 判断是否为长包模式
                    if all(x == '00' for x in ls[9:]):  # 检查第9位开始是否全为0，判定为长包
                        self.ClusterDatadistinguish(ls)
                        
                    elif all(x == ls[9] for x in ls[10:]):    #检查是否为应答包
                        InstrCRC = ls[8] + ls[9]  # 取出数据包中的CRC码
    
                        CRCVerify = ls[2] + ls[3] + ls[4] + ls[5] + ls[6] + ls[7]  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.Insdistinguish(ls)
                    else:   # 若不是应答包也不是短包，则是长包，长包数据应答包长度为138
                        InstrCRC = ls[136] + ls[137]

                        CRCVerify = ''.join(ls[2:136])  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.ClusterDatadistinguish(ls)
                elif AcquireMode == "事例击中":  # 判断是否为长包模式
                    if all(x == '00' for x in ls[9:]):  # 检查第9位开始是否全为0，判定为长包
                        self.HitDatadistinguish(ls)
                        
                    elif all(x == ls[9] for x in ls[10:]):    #检查是否为应答包
                        InstrCRC = ls[8] + ls[9]  # 取出数据包中的CRC码
    
                        CRCVerify = ls[2] + ls[3] + ls[4] + ls[5] + ls[6] + ls[7]  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.Insdistinguish(ls)
                    else:   # 若不是应答包也不是短包，则是长包，长包数据应答包长度为138
                        InstrCRC = ls[72] + ls[73]

                        CRCVerify = ''.join(ls[2:72])  # 取出数据包中的CRC校验字段
                        CRCVerify_bytes = bytes.fromhex(CRCVerify)  # 将CRC校验字段转化为字节
                        CRCVerify_result = self.InstrCRCverify(CRCVerify_bytes)  # 对数据包进行CRC校验
                        CRCVerify_result = format(CRCVerify_result, '04X')  # 将CRC校验结果转化为十六进制
                        self.HitDatadistinguish(ls)
                        
            elif ls[0] == '11' and ls[1] == '11' and ls[2] == '11' and ls[3] == '11' and ls[4] == '11':
                self.ifSynCtrlTriggerSuccess = 0
                self.ifSynCtrlTriggerSuccess = self.SynCtrlTriggerCommandJudge(ls)
                
            elif ls[0] == '00' and ls[1] == '00' and ls[2] == '00' and ls[3] == '00' and ls[4] == '00':
                self.CommunicationTextBrowser.append("中间板未能收到正确的同步触发应答信号，请再次触发或停机检查")
                
            else:
                self.CommunicationTextBrowser.append("应答包头错误")
            
        else:
            self.CommunicationTextBrowser.append("数据长度不正确")

    def AckCodeVerify(self, data):  # 应答码校验
        AckCode = data[5]  # 取出应答包中的应答码

        if AckCode == 'F1':
            self.CommunicationTextBrowser.append("工作正常")
            self.SpectroscopyTextBrowser.append('工作正常')
            return 0000
        elif AckCode == 'F2':
            self.SpectroscopyTextBrowser.append('指令帧序号不连续，疑似出现指令丢失，建议检查')
            self.CommunicationTextBrowser.append("指令帧序号不连续，疑似出现指令丢失，建议检查")
            return 0000
        elif AckCode == 'F3':
            self.SpectroscopyTextBrowser.append('探测器正处于Busy状态，不执行指令，建议等待后再试')
            self.CommunicationTextBrowser.append("探测器正处于Busy状态，不执行指令，建议等待后再试")
            # EXTENSION
            # return 1111
            return 0000
        elif AckCode == 'F4':
            # EXTENSION
            # self.SpectroscopyTextBrowser.append('CRC校验出错，指令包可能已损坏，不执行指令，建议检查')
            # self.CommunicationTextBrowser.append("CRC校验出错，指令包可能已损坏，不执行指令，建议检查")
            # return 1111
            return 0000
        elif AckCode == 'F5':
            self.SpectroscopyTextBrowser.append('无效指令：指令包中的指令码无法识别，不执行指令，建议检查')
            self.CommunicationTextBrowser.append("无效指令：指令包中的指令码无法识别，不执行指令，建议检查")
            return 1111
        else:
            self.SpectroscopyTextBrowser.append('未定义应答码，建议检查')
            self.CommunicationTextBrowser.append("未定义应答码，建议检查")
            return 1111

    def Insdistinguish(self, data):  # 应答序号与指令序号匹配，进一步区分指令
        ls = data

        VerifyCode = self.AckCodeVerify(ls)  # 应答码检测
        if VerifyCode == 0000:
            pass
        else:
            return

        self.DeviceID = int(ls[4], 16)  # 取出应答包中的设备ID
        self.AckSN = int(ls[2]+ls[3], 16)  # 取出应答包中的应答序号
        self.AckInsdistinguish(ls)
        # EXTENSION
        # if self.AckSN == self.InsSN - 1:  # 匹配应答指令序号与指令帧序号（指令帧序号每次发送后会+1）
        #     self.SpectroscopyTextBrowser.append('应答序号与指令序号匹配')
        #     self.CommunicationTextBrowser.append('应答序号与指令序号匹配')
        #     self.AckInsdistinguish(ls)  # 根据应答参数做下一步行为
        # else:
        #     self.SpectroscopyTextBrowser.append('应答序号与指令序号不匹配，疑似出现指令丢失，建议检查')
        #     self.CommunicationTextBrowser.append('应答序号与指令序号不匹配，疑似出现指令丢失，建议检查')

    def AckInsdistinguish(self, data):  # 应答指令区分
        Ackls = data
        str_Ackls = " ".join(Ackls)  # 将应答指令组合

        Instrls = []  # 用于存储上一条发送的指令
        for i in range(len(self.Instruction) // 2):  # 对十六进制数据进行byte解码，且转化为大写
            hex_value = self.Instruction[i * 2].upper() + self.Instruction[i * 2 + 1].upper()
            Instrls.append(hex_value)

        if Instrls[5] == '01':  # 命令码为0x01
            if Instrls[6] == '00':  # 指令回环测试 
                self.CommunicationTextBrowser.append('回复指令：%s' % Ackls[7])

            elif Instrls[6] + Instrls[7] == '0200':  # 读取控制寄存器的配置
                CtrlReg = int(Ackls[6] + Ackls[7], 16)  # 取出应答包中的控制寄存器
                self.CommunicationTextBrowser.append('控制寄存器配置为：%s' % f"{CtrlReg:016b}")
                CtrlRegls = []
                for i in range(16):
                    CtrlRegls.append(f"{CtrlReg:016b}"[i])

                if CtrlRegls[11] == '0':  # 测试信号输出使能
                    self.CommunicationTextBrowser.append('测试信号设置：输出禁止')
                elif CtrlRegls[11] == '1':
                    self.CommunicationTextBrowser.append('测试信号设置：输出使能')

                if CtrlRegls[12] + CtrlRegls[13] == '00':  # 数据输出模式
                    self.CommunicationTextBrowser.append('数据输出模式：短数据包模式')
                elif CtrlRegls[12] + CtrlRegls[13] == '01':
                    self.CommunicationTextBrowser.append('数据输出模式：长数据包模式')

                if CtrlRegls[14] == '0':  # 工作模式设置
                    self.CommunicationTextBrowser.append('工作模式设置：正常取数模式')
                elif CtrlRegls[14] == '1':
                    self.CommunicationTextBrowser.append('工作模式设置：电子学刻度模式')

                if CtrlRegls[15] == '0':  # 触发接收使能
                    self.CommunicationTextBrowser.append('触发接收使能：禁止')
                elif CtrlRegls[15] == '1':
                    self.CommunicationTextBrowser.append('触发接收使能：使能')

            elif Instrls[6] + Instrls[7] == '0400':  # 读取阈值配置
                Threshold = int(Ackls[6] + Ackls[7], 16)
                self.CommunicationTextBrowser.append('阈值配置为：%d' % Threshold)
                self.SpectroscopyTextBrowser.append('阈值配置为：%d' % Threshold)
                self.ThresholdLineEdit.setText(str(Threshold))

            elif Instrls[6] + Instrls[7] == '0600':
                AcquireTime = int(Ackls[6] + Ackls[7], 16)
                self.CommunicationTextBrowser.append('采集时间配置为：%d' % AcquireTime)
                self.SpectroscopyTextBrowser.append('采集时间配置为：%d' % AcquireTime)
                self.AcquireTimeLineEdit.setText(str(AcquireTime))

        else:
            pass

    def ShortDatadistinguish(self, data):  # 短包数据区分

        LS_ShortData = data

        self.AckSN = int(LS_ShortData[2]+LS_ShortData[3], 16)  # 取出应答包中的应答序号
        self.DeviceID = LS_ShortData[4]  # 取出应答包中的设备ID

        VerifyCode = self.AckCodeVerify(LS_ShortData)  # 应答码检测
        if VerifyCode == 0000:  # 应答码检测通过
            pass
        else:
            return
        self.ShortAckInsdistinguish(LS_ShortData)
        # EXTENSION
        # if self.AckSN == self.InsSN - 1:  # 匹配应答指令序号与指令帧序号（指令帧序号每次发送后会+1）
        #     self.SpectroscopyTextBrowser.append('应答序号与指令序号匹配')
        #     self.CommunicationTextBrowser.append('应答序号与指令序号匹配')
        #     self.ShortAckInsdistinguish(LS_ShortData)
        # else:
        #     self.SpectroscopyTextBrowser.append('应答序号与指令序号不匹配，疑似出现指令丢失，建议检查')
        #     self.CommunicationTextBrowser.append('应答序号与指令序号不匹配，疑似出现指令丢失，建议检查')

    def ShortAckInsdistinguish(self, data):  # 短数据包数据接收，并绘制能谱图

        LS_Shortdata = data  # 获取长包数据
        self.ChannelLongDATA = {}  # 建立一个空字典，用于存储通道数据

        # TriggerChSN = int(LS_Longdata[6], 16)  # 获取触发通道号
        # Length_Longdata = int(LS_Longdata[7], 16)  # 获取长包数据长度

        for i in range(32):  # 遍历32个通道
            start_index = 8 + i * 2  # 计算每个键对应的切片开始索引
            end_index = start_index + 2  # 切片结束索引不包括在内，所以+4
            hex_str = ''.join(LS_Shortdata[start_index:end_index])  # 从LS中提取四个元素，并转化为一个字符串

            value = int(hex_str, 16)  # 将合并后的十六进制字符串转化为整数

            # 将键和转换后的值添加到字典中
            self.ChannelLongDATA[f'CH{i}'] = value

        self.spectrum_data_list.append(list(self.ChannelLongDATA.values())) # 将数据添加到列表中
        
        self.plotCanvas = MyMatplotlibFigure(width=7.1, height=5.51, dpi=1200)

        pixmap = self.plotCanvas.draw_bar_chart(self.ChannelLongDATA)  # 绘制能谱图

        self.scene.clear()  # 清除场景中的所有项
        self.scene.addPixmap(pixmap)  # 添加新的 QPixmap 到场景
        self.SpectroscopyView.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)  # 适应图形到视图

    def LongDatadistinguish(self, data):  # 长包数据区分
        LS_LongData = data

        VerifyCode = self.AckCodeVerify(LS_LongData)  # 应答码检测
        if VerifyCode == 0000:  # 应答码检测通过
            pass
        else:
            return

        self.AckSN = int(LS_LongData[2]+LS_LongData[3], 16)  # 取出应答包中的应答序号
        self.DeviceID = LS_LongData[4]  # 取出应答包中的设备ID
        self.LongAckInsdistinguish(LS_LongData)
        # EXTENSION
        # if self.AckSN == self.InsSN - 1:  # 匹配应答指令序号与指令帧序号（指令帧序号每次发送后会+1）
        #     self.SpectroscopyTextBrowser.append('应答序号与指令序号匹配')
        #     self.CommunicationTextBrowser.append('应答序号与指令序号匹配')
        #     self.LongAckInsdistinguish(LS_LongData)
        # else:
        #     self.SpectroscopyTextBrowser.append('应答序号与指令序号不匹配，疑似出现指令丢失，建议检查')
        #     self.CommunicationTextBrowser.append('应答序号与指令序号不匹配，疑似出现指令丢失，建议检查')

    def LongAckInsdistinguish(self, data):  # 长数据包数据接收，并绘制能谱图

        LS_Longdata = data  # 获取长包数据
        self.ChannelLongDATA = {}  # 建立一个空字典，用于存储通道数据

        # TriggerChSN = int(LS_Longdata[6], 16)  # 获取触发通道号
        # Length_Longdata = int(LS_Longdata[7], 16)  # 获取长包数据长度

        for i in range(32):  # 遍历32个通道
            start_index = 8 + i * 4  # 计算每个键对应的切片开始索引
            end_index = start_index + 4  # 切片结束索引不包括在内，所以+4
            hex_str = ''.join(LS_Longdata[start_index:end_index])  # 从LS中提取四个元素，并转化为一个字符串

            value = int(hex_str, 16)  # 将合并后的十六进制字符串转化为整数

            # 将键和转换后的值添加到字典中
            self.ChannelLongDATA[f'CH{i}'] = value
        
        self.spectrum_data_list.append(list(self.ChannelLongDATA.values())) # 将数据添加到列表中

        self.plotCanvas = MyMatplotlibFigure(width=7.1, height=5.51, dpi=1200)

        pixmap = self.plotCanvas.draw_bar_chart(self.ChannelLongDATA)  # 绘制能谱图

        self.scene.clear()  # 清除场景中的所有项
        self.scene.addPixmap(pixmap)  # 添加新的 QPixmap 到场景
        self.SpectroscopyView.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)  # 适应图形到视图

    def ClusterDatadistinguish(self, data):  # 长包数据区分
        LS_LongData = data

        VerifyCode = self.AckCodeVerify(LS_LongData)  # 应答码检测
        if VerifyCode == 0000:  # 应答码检测通过
            pass
        else:
            return

        self.AckSN = int(LS_LongData[2]+LS_LongData[3], 16)  # 取出应答包中的应答序号
        self.DeviceID = LS_LongData[4]  # 取出应答包中的设备ID
        self.ClusterAckInsdistinguish(LS_LongData)

    def ClusterAckInsdistinguish(self, data):  # 长数据包数据接收，并绘制能谱图

        LS_Longdata = data  # 获取长包数据
        self.ChannelLongDATA = {}  # 建立一个空字典，用于存储通道数据

        for i in range(63):  # 遍历32个通道
            start_index = 8 + i * 2  # 计算每个键对应的切片开始索引
            end_index = start_index + 2  # 切片结束索引不包括在内，所以+4
            hex_str = ''.join(LS_Longdata[start_index:end_index])  # 从LS中提取四个元素，并转化为一个字符串

            value = int(hex_str, 16)  # 将合并后的十六进制字符串转化为整数

            # 将键和转换后的值添加到字典中
            self.ChannelLongDATA[f'CH{i}'] = value
        
        self.spectrum_data_list.append(list(self.ChannelLongDATA.values())) # 将数据添加到列表中

        self.plotCanvas = MyMatplotlibFigure(width=7.1, height=5.51, dpi=1200)
        self.plotCanvas.axes_channels = 63

        pixmap = self.plotCanvas.draw_bar_chart(self.ChannelLongDATA)  # 绘制能谱图

        self.scene.clear()  # 清除场景中的所有项
        self.scene.addPixmap(pixmap)  # 添加新的 QPixmap 到场景
        self.SpectroscopyView.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)  # 适应图形到视图

    def HitDatadistinguish(self, data):  # 长包数据区分
        LS_LongData = data

        VerifyCode = self.AckCodeVerify(LS_LongData)  # 应答码检测
        if VerifyCode == 0000:  # 应答码检测通过
            pass
        else:
            return

        self.AckSN = int(LS_LongData[2]+LS_LongData[3], 16)  # 取出应答包中的应答序号
        self.DeviceID = LS_LongData[4]  # 取出应答包中的设备ID
        self.HitAckInsdistinguish(LS_LongData)

    def HitAckInsdistinguish(self, data):  # 长数据包数据接收，并绘制能谱图

        LS_Longdata = data  # 获取长包数据
        
        self.ChannelLongDATA = {}  # 建立一个空字典，用于存储通道数据

        for i in range(32):
            self.ChannelLongDATA[f"CH{i}"] = 0
        for i in range(16):
            start_index = 8 + i * 4  # 计算每个键对应的切片开始索引
            end_index = start_index + 4  # 切片结束索引不包括在内，所以+4
            #hit_channels_bits = int8toint32(LS_Longdata[start_index:end_index])
            #print(hex(hit_channels_bits))
            hex_str = ''.join(LS_Longdata[start_index:end_index])  # 从LS中提取四个元素，并转化为一个字符串
            value = int(hex_str, 16)  # 将合并后的十六进制字符串转化为整数
            for j in range(32):
                bit = (value>>j)&0x1
                self.ChannelLongDATA[f'CH{j}'] += bit
        self.spectrum_data_list.append(list(self.ChannelLongDATA.values())) # 将数据添加到列表中

        self.plotCanvas = MyMatplotlibFigure(width=7.1, height=5.51, dpi=1200)

        pixmap = self.plotCanvas.draw_bar_chart(self.ChannelLongDATA)  # 绘制能谱图

        self.scene.clear()  # 清除场景中的所有项
        self.scene.addPixmap(pixmap)  # 添加新的 QPixmap 到场景
        self.SpectroscopyView.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)  # 适应图形到视图
    def resizeEvent(self, event):  # 当窗口大小改变时，重新适配视图   目前好像没什么用
        super().resizeEvent(event)
        if self.scene.items():  # 检查场景中是否有项目
            self.SpectroscopyView.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)  # 重新适配到视图

    def on_InsLoopTest(self):  # 指令回环测试
        InstrLoopTest = 0X0100  # 指令回环测试的默认字段
        str_InstrLoopTest = f"{InstrLoopTest:04X}"

        str_InstrLoop = self.InstrLooplineEdit.text()  # 获取指令回环测试的指令
        self.send_data(self.InstrCombination(str_InstrLoopTest, str_InstrLoop))  # 发送指令回环测试的指令

    def InstrCombination(self, *args):  # 将指令组合，传入的参数为十六进制字符串

        concatenated_string = "".join(args)  # 将传入的参数先进行组合

        # 先进行CRC校验
        crc_InstrCombination = self.InstrSN + self.str_InstrReserve + concatenated_string  # 连接需要进行CRC校验的字符串
        crc_int_InstrCombination = int(crc_InstrCombination, 16)  # 将字符串转化为十六进制整数

        crc_bytes_InstrCombination = crc_int_InstrCombination.to_bytes(6, byteorder='big')  # 将整数转化为字节
        InsCRC = self.InstrCRCverify(crc_bytes_InstrCombination)  # CRC校验
        str_InsCRC = f"{InsCRC:04X}"

        # 完整指令组合
        str_InstrCombination = self.str_InstrPacketStart + self.InstrSN + self.str_InstrReserve + concatenated_string + str_InsCRC
        InstrCombination = int(str_InstrCombination, 16)
        InstrCombinationresult = InstrCombination.to_bytes(10, byteorder='big')  # 将字符串转化为字节
        return InstrCombinationresult

    def InstrCRCverify(self, data_bytes, poly=0x1021, init_val=0xFFFF):  # 指令CRC校验

        crc = init_val
        for byte in data_bytes:
            crc ^= byte << 8
            for _ in range(8):  # 处理每一位
                if (crc & 0x8000):  # 判断最高位是否为1
                    crc = (crc << 1) ^ poly
                else:
                    crc <<= 1
                crc &= 0xFFFF  # 保证CRC值为16位
        return crc

    def openFolderDialog(self):  # 选择文件路径，并将文件路径填充在文本框中
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        folderPath = QFileDialog.getExistingDirectory(
            None,
            "Select Data Folder",
            "",
            options=options)
        if folderPath:
            self.FilePathLineEdit.setText(folderPath)

    def UDPBind(self):  # 绑定UDP
        self.udpSocket = QUdpSocket(self)  # 创建一个UDP套接字
        self.localPort = 8081  # 本地端口
        self.udpSocket.bind(QHostAddress.Any, self.localPort)  # 绑定本地地址,自动获取本地所有ip地址(IPV4)

        self.receiverThread = ReceiverThread(self.udpSocket)   # 创建子线程，用于监听到来的UDP数据
        self.receiverThread.dataReceived.connect(self.onDataReceived)
        self.receiverThread.ReceivedError.connect(self.CommunicationTextBrowser.append)  # 将错误信息显示在通信文本框中
        self.receiverThread.start()

        self.IPAddress = self.IPAddressLineEdit.text()
        self.UDPPort = int(self.UDPServerLineEdit.text())
        self.Server_addr = (self.IPAddress, self.UDPPort)  # 服务端地址
        self.SpectroscopyTextBrowser.append("UDP 绑定地址为 %s:%d" % (self.IPAddress, self.UDPPort))

    def SendIns(self):  # 发送指令，只需要输入命令码和命令参数，也就是六位十六进制数
        message = self.SendInsLineEdit.text()
        
        try:
            # 尝试将十六进制字符串转换为 bytes 型变量
            message_bytes = bytes.fromhex(message)
            self.send_data(message_bytes)  # 发送指令
        except ValueError:
            # 如果转换失败，显示错误消息
            self.CommunicationTextBrowser.append("无效的十六进制字符串")
        except Exception as e:
            print("Error: ", e)

    def AcquireData(self):  # 采集数据
        InsAcquireData = 0X020100  # 采集数据的命令码与命令参数
        str_InsAcquireData = f"{InsAcquireData:06X}"
        Instruction_AcquireData = self.InstrCombination(str_InsAcquireData)  # 组合指令
        self.send_data(Instruction_AcquireData)  # 发送指令

    def CountsRest(self):  # 计数复位
        InsCountsRest = 0X030100  # 采集数据的命令码与命令参数
        str_InsCountsRest = f"{InsCountsRest:06X}"
        Instruction_CountsRest = self.InstrCombination(str_InsCountsRest)  # 组合指令
        self.send_data(Instruction_CountsRest)  # 发送指令

    def ThresholdConfig(self, dac_value):  # 阈值配置
        if dac_value == 0:
            ThresholdStr = self.ThresholdLineEdit.text()  # 获取阈值
            ThresholdValue = int(ThresholdStr)  # 将阈值转化为整数
        else:
            ThresholdValue = dac_value
        
        if ThresholdValue > 4095:
            self.SpectroscopyTextBrowser.append("阈值超出范围")
            self.CommunicationTextBrowser.append("阈值超出范围")
            return

        ThresholdConfigCode = 0X05  # 阈值配置的命令码
        str_ThresholdConfigCode = f"{ThresholdConfigCode:02X}"

        str_ThresholdValue = f"{ThresholdValue:04X}"

        InsThresholdConfig = self.InstrCombination(str_ThresholdConfigCode, str_ThresholdValue)  # 组合指令
        self.send_data(InsThresholdConfig)  # 发送指令

    def AcquireConfig(self):  # 采集配置
        AcquireTime = self.AcquireTimeLineEdit.text()  # 获取采集时长
        AcquireTimeValue = int(AcquireTime)  # 将阈值转化为整数

        if AcquireTimeValue > 65535:
            self.SpectroscopyTextBrowser.append("采集时间超出范围")
            self.CommunicationTextBrowser.append("采集时间超出范围")
            return

        AcquireTimeConfigCode = 0X06  # 阈值配置的命令码
        str_AcquireTimeConfigCode = f"{AcquireTimeConfigCode:02X}"

        str_AcquireTime = f"{AcquireTimeValue:04X}"
        InsAcquireTime = self.InstrCombination(str_AcquireTimeConfigCode, str_AcquireTime)  # 组合指令
        self.send_data(InsAcquireTime)  # 发送指令

    def on_AcquireMode_changed(self):  # 采集模式改变时，通过配置按钮获取采集模式
        AcquireMode = self.AcquireModeBox.currentText()
        if AcquireMode == "短包":
            self.CtrlRegDataMode = 0B00
        elif AcquireMode == "长包":
            self.CtrlRegDataMode = 0B01
        elif AcquireMode == "簇团计数":
            self.CtrlRegDataMode = 0B10
        elif AcquireMode == "事例击中":
            self.CtrlRegDataMode = 0B11    

        ctrl_reg = self.CtrlRegParam_combined()

        Instr_AcquireMode = self.InstrCombination(ctrl_reg)

        self.send_data(Instr_AcquireMode)

    def CtrlRegRead(self):  # 控制寄存器读取
        InsCtrlRegRead = 0X010200
        str_InsCtrlRegRead = f"{InsCtrlRegRead:06X}"
        Instruction_CtrlRegRead = self.InstrCombination(str_InsCtrlRegRead)
        self.send_data(Instruction_CtrlRegRead)

    def FPGAReset(self):  # FPGA复位
        InsFPGAReset = 0X010100
        str_InsFPGAReset = f"{InsFPGAReset:06X}"
        Instruction_FPGAReset = self.InstrCombination(str_InsFPGAReset)
        self.send_data(Instruction_FPGAReset)
        
    def FileSave(self):  # 文件保存
        FilePath = self.FilePathLineEdit.text()
        FileName = self.FileNameLineEdit.text()

        if not FileName.endswith('.txt'):
            FileName += '.txt'

        full_path = os.path.join(FilePath, FileName)

        try:
            # 检查并创建文件所在的目录
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
        except Exception as e:
            self.SpectroscopyTextBrowser.append("文件路径错误,无法创建：%s" % e)
            return

        # 写入文件，如果文件不存在，自动创建
        with open(full_path, 'w', encoding='utf-8') as file:
            for key, value in self.ChannelLongDATA.items():
                file.write(f"{value}\n")

        self.SpectroscopyTextBrowser.append("文件保存成功")

    def FileDelete(self):  # 文件删除
        FilePath = self.FilePathLineEdit.text()
        FileName = self.FileNameLineEdit.text()
        if not FileName.endswith('.txt'):
            FileName += '.txt'
            
        full_path = os.path.join(FilePath, FileName)

        if os.path.exists(full_path):
            os.remove(full_path)
            self.SpectroscopyTextBrowser.append(f"文件已删除：{full_path}")
        else:
            self.SpectroscopyTextBrowser.append(f"文件不存在：{full_path}")

    def CtrlRegConfig(self):   # 配置控制寄存器，包括触发接收使能，测试信号使能，工作模式选择，修改测试通道和成形时间
        TestChannel = self.TestChannelBox.currentText()
        if TestChannel == "ch1":
            self.CtrlRegTestChannel = 0B00
        elif TestChannel == "ch9":
            self.CtrlRegTestChannel = 0B01
        elif TestChannel == "ch17":
            self.CtrlRegTestChannel = 0B10
        elif TestChannel == "ch25":
            self.CtrlRegTestChannel = 0B11
        
        FormingTime = self.FormingTimeBox.currentText()
        if FormingTime == "40ns":
            self.CtrlRegFormingTime = 0B00
        elif FormingTime == "80ns":
            self.CtrlRegFormingTime = 0B01
        elif FormingTime == "120ns":
            self.CtrlRegFormingTime = 0B10
        elif FormingTime == "160ns":
            self.CtrlRegFormingTime = 0B11
        
        SerialPortRate = self.SerialPortRateBox.currentText()
        if SerialPortRate == "Debug模式":
            self.CtrlRegSerialPortRate = 0B00_00
        elif SerialPortRate == "正常":
            self.CtrlRegSerialPortRate = 0B00_01
        
        TestSignalEnableMode = self.TestSignalEnableBox.currentText()
        if TestSignalEnableMode == "禁止使能":
            self.CtrlRegTestSignalOutputEnable = 0B00_00
        elif TestSignalEnableMode == "使能":
            self.CtrlRegTestSignalOutputEnable = 0B00_01    
        
        WorkingMode = self.WorkingModeBox.currentText()
        TriggerReceiveEnable = self.TriggerReceiveEnableBox.currentText()
        if WorkingMode == "正常取数" and TriggerReceiveEnable == "禁止使能":
            self.CtrlRegWorkingMode = 0B00
        elif WorkingMode == "电子学刻度" and TriggerReceiveEnable == "禁止使能":
            self.CtrlRegWorkingMode = 0B10
        elif WorkingMode == "正常取数" and TriggerReceiveEnable == "使能":      
            self.CtrlRegWorkingMode = 0B01
        elif WorkingMode == "电子学刻度" and TriggerReceiveEnable == "使能":  
            self.CtrlRegWorkingMode = 0B11
        
            
        ctrl_reg = self.CtrlRegParam_combined()
        
        Instr_CtrlRegConfig = self.InstrCombination(ctrl_reg)
        
        self.send_data(Instr_CtrlRegConfig)
    
    def SynCtrlTrigger(self):   # 同步控制触发，EB 90 00 00 00 0C AE 00 00 00
        # 两种方式二选一 这个是有指令帧的
        # InsSynCtrlTrigger = 0X0AAE00  # 采集数据的命令码与命令参数
        # str_InsSynCtrlTrigger = f"{InsSynCtrlTrigger:06X}"
        # Instruction_SynCtrlTrigger = self.InstrCombination(str_InsSynCtrlTrigger)  # 组合指令
        # self.send_data(Instruction_SynCtrlTrigger)  # 发送指令
        
        self.InsSN -= 1 # 恢复指令帧序号
        InsSynCtrlTrigger = 0XEB900000000CAE000000
        str_InsSynCtrlTrigger = f"{InsSynCtrlTrigger:20X}"
        Instruction_SynCtrlTrigger = bytes.fromhex(str_InsSynCtrlTrigger)
        self.send_data(Instruction_SynCtrlTrigger)  # 发送指令
    
    def SynCtrlTriggerCommandJudge(self, data): # 同步触发指令应答判断
        ls = data
        Flag = 0
        if ls[5] == '00' and ls[6] == '00' and ls[7] == '00' and ls[8] == '00' and ls[9] == '00':
            Flag = 1
        elif ls[5] == '22' and ls[6] == '22' and ls[7] == '22' and ls[8] == '22' and ls[9] == '22':
            Flag = 2
        else: 
            self.CommunicationTextBrowser.append("同步触发指令应答异常")
        
        return Flag
        
    def CtrlRegParam_combined(self):    # 如果需要配置控制寄存器的值, 将控制寄存器的多位二进制字面量进行组合，再与命令码进行组合
        CtrlRegParam_combined = (self.CtrlRegTestChannel << 14) | (self.CtrlRegFormingTime << 12) | (self.CtrlRegSerialPortRate << 8) | (self.CtrlRegTestSignalOutputEnable << 4) | (self.CtrlRegDataMode << 2) | self.CtrlRegWorkingMode  # 将控制寄存器参数组合
        str_CtrlRegParam = f"{CtrlRegParam_combined:04X}"  # 将组合后的参数转化为十六进制字符串
        InstrCode = 0X04  # 配置控制寄存器的命令码
        str_InstrCode = f"{InstrCode:02X}"
        str_CtrlReg = str_InstrCode + str_CtrlRegParam  # 三字节的十六进制字符串
        return str_CtrlReg
    
    def ScaleThreshold(self):   # 配置刻度阈值，配置刻度DAC
        ScaleThresholdStr = self.ScaleThresholdEdit.text()  
        ScaleThresholdValue = int(ScaleThresholdStr)  

        if ScaleThresholdValue > 4095:
            self.SpectroscopyTextBrowser.append("刻度超出范围")
            self.CommunicationTextBrowser.append("刻度超出范围")
            return

        ScaleThresholdConfigCode = 0X07  
        str_ScaleThresholdConfigCode = f"{ScaleThresholdConfigCode:02X}"

        str_ScaleThresholdValue = f"{ScaleThresholdValue:04X}"

        InsScaleThresholdConfig = self.InstrCombination(str_ScaleThresholdConfigCode, str_ScaleThresholdValue)  # 组合指令
        self.send_data(InsScaleThresholdConfig)  # 发送指令
    
    def ScaleTimeInterval(self):    # 配置刻度时间间隔
        ScaleTimeInterval = self.ScaleTimeIntervalEdit.text()  
        ScaleTimeIntervalValue = int(ScaleTimeInterval)  

        if ScaleTimeIntervalValue > 65535:
            self.SpectroscopyTextBrowser.append("采集时间超出范围")
            self.CommunicationTextBrowser.append("采集时间超出范围")
            return

        ScaleTimeIntervalConfigCode = 0X08  
        str_ScaleTimeIntervalConfigCode = f"{ScaleTimeIntervalConfigCode:02X}"

        str_ScaleTimeInterval = f"{ScaleTimeIntervalValue:04X}"

        InsScaleTimeInterval = self.InstrCombination(str_ScaleTimeIntervalConfigCode, str_ScaleTimeInterval)  # 组合指令
        self.send_data(InsScaleTimeInterval)  # 发送指令
    
    def SaveConfig(self):   # 保存配置至当前文件夹
        parameters = {
            "Threshold": self.ThresholdLineEdit.text(),
            "AcquireTime": self.AcquireTimeLineEdit.text(),
            "AcquireMode": self.AcquireModeBox.currentText(),
            "ScaleThresholdStr": self.ScaleThresholdEdit.text(),
            "ScaleTimeInterval": self.ScaleTimeIntervalEdit.text(),
            "TestChannel": self.TestChannelBox.currentText(),
            "FormingTime": self.FormingTimeBox.currentText(),
            "TriggerReceiveEnable": self.TriggerReceiveEnableBox.currentText(),
            "TestSignalEnableMode": self.TestSignalEnableBox.currentText(),
            "WorkingMode": self.WorkingModeBox.currentText(),
            "SerialPortRate": self.SerialPortRateBox.currentText(),
            "FilePath": self.Data_transmission.FilePathLineEdit.text(),
            # 要保存的配置参数
        }
        
        current_directory = os.getcwd() # 获取当前工作目录
        config_directory = os.path.join(current_directory, "Config")
        
        if not os.path.exists(config_directory):    # 如果没有Config文件夹则创建一个
            os.makedirs(config_directory)
            
        # 弹出保存文件对话框，让用户选择文件名和位置
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Parameters", os.path.join(config_directory, "配置.json"), "JSON Files (*.json);;All Files (*)", options=options)

        if file_path:
            with open(file_path, 'w') as file:
                json.dump(parameters, file)
            self.CommunicationTextBrowser.append(f"配置文件保存至 {file_path}")
            
    def LoadConfig(self):   # 导入配置文件，自动改变参数
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(self, "Load Parameters", "", "JSON Files (*.json);;All Files (*)", options=options)

        if file_path:
            with open(file_path, 'r') as file:
                parameters = json.load(file)

            self.ThresholdLineEdit.setText(parameters.get("Threshold", ""))
            self.AcquireTimeLineEdit.setText(parameters.get("AcquireTime", ""))
            self.AcquireModeBox.setCurrentText(parameters.get("AcquireMode", ""))
            self.ScaleThresholdEdit.setText(parameters.get("ScaleThresholdStr", ""))
            self.ScaleTimeIntervalEdit.setText(parameters.get("ScaleTimeInterval", ""))
            self.TestChannelBox.setCurrentText(parameters.get("TestChannel", ""))
            self.FormingTimeBox.setCurrentText(parameters.get("FormingTime", ""))
            self.TriggerReceiveEnableBox.setCurrentText(parameters.get("TriggerReceiveEnable", ""))
            self.TestSignalEnableBox.setCurrentText(parameters.get("TestSignalEnableMode", ""))
            self.WorkingModeBox.setCurrentText(parameters.get("WorkingMode", ""))
            self.SerialPortRateBox.setCurrentText(parameters.get("SerialPortRate", ""))
            self.FilePathLineEdit.setCurrentText(parameters.get("FilePath", ""))
            # 设置其他加载的参数
            print(f"从 {file_path} 导入配置文件")
            self.CommunicationTextBrowser.append(f"从 {file_path} 导入配置文件")
        else:
            print("未选中配置文件")
            self.CommunicationTextBrowser.append("未选中配置文件")        
            
    def ClearSpec(self):    # 清除能谱图
        self.plotCanvas.clear_plot()  # 清除绘图
        self.scene.clear()  # 清除场景中的所有项
        self.scene.addWidget(self.plotCanvas)  # 重新添加绘图组件到场景中
        self.SpectroscopyView.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatioByExpanding)  # 适应图形到视图
        self.CommunicationTextBrowser.append("能谱图已清除")
        self.SpectroscopyTextBrowser.append("能谱图已清除")
        
    def SampStepTotalLength(self):  # 设置测量S曲线前的采集步长与总长度
        dialog = SampStepTotalLengthDialog()
        if dialog.exec_() == QDialog.Accepted:
            self.sampling_step, self.total_length = dialog.getInputs()  # 得到对话框中的采集步长和总长度
            # 将浮点数转化为整数
            self.sampling_step = int(self.sampling_step)    
            self.total_length = int(self.total_length)
            self.CommunicationTextBrowser.append(f'采集步长: {self.sampling_step}, 采集总长度: {self.total_length}')
            self.SpectroscopyTextBrowser.append(f'采集步长: {self.sampling_step}, 采集总长度: {self.total_length}')
            
    def SCurve(self):   # 一键测量S曲线
        self.SCurveHandler.measure_s_curve(self.sampling_step, self.total_length)

    def Loop(self): # 延迟应答判定函数
        self.ifdatareceived = 0
        self.loop_result = 0
        
        loop = QEventLoop()     # 创建一个循环
        
        def check_data_received():
            if self.ifdatareceived == 1:
                self.loop_result = 1
            loop.quit()
        
        QTimer.singleShot(1000, check_data_received)  # 设置1s的定时器，触发时退出循环
        loop.exec_()    # 开始循环并等待直到定时器触发退出循环
        
        return self.loop_result
    
    def retry_loop(self, action, success_message, fail_message):
        success = False
        while not success:
            action()
            if self.Loop() == 1:
                success = True
                self.SpectroscopyTextBrowser.append(success_message)
                self.CommunicationTextBrowser.append(success_message)
            else:
                self.SpectroscopyTextBrowser.append(fail_message)
                self.CommunicationTextBrowser.append(fail_message)        
    
    def PeriodCollect(self):    # 周期同步触发和读数
        PeriodCollect = self.PeriodEdit.text()
        PeriodCollectValue = int(PeriodCollect)
        
        all_data = []  # 存储所有周期的数据
                
        for i in range(PeriodCollectValue):
            self.retry_loop(self.CountsRest, 
                    "清零成功", 
                    "清零信号未接收到应答包，重新发送")
            
            self.ifSynCtrlTriggerSuccess = 0
            self.SynCtrlTrigger()
            
            AcquireTime = self.AcquireTimeLineEdit.text()  # 获取采集时长
            AcquireTimeValue = int(AcquireTime)  # 将阈值转化为整数
            Delay = int(AcquireTimeValue/10 +2000) # 1倍采集时长+2s，定时器的单位是1ms
            loop = QEventLoop()     # 创建一个循环
            QTimer.singleShot(Delay, loop.quit)  # 设置定时器，触发时退出循环
            loop.exec_()    # 开始循环并等待直到定时器触发退出循环
            
            if self.ifSynCtrlTriggerSuccess == 2:
                self.retry_loop(self.AcquireData, 
                    "读数成功", 
                    "读数信号未接收到应答包，重新发送")
                all_data.append(self.get_channel_data())
            elif self.ifSynCtrlTriggerSuccess == 1:
                self.CommunicationTextBrowser.append(f"周期采集中第{i+1}次采集探测器能谱采集工作异常，建议检查")
                self.SpectroscopyTextBrowser.append(f"周期采集中第{i+1}次采集探测器能谱采集工作异常，建议检查")
                break
            elif self.ifSynCtrlTriggerSuccess == 0:
                self.SynCtrlTrigger()
                QTimer.singleShot(Delay, loop.quit)  # 设置1s的定时器，触发时退出循环
                loop.exec_()    # 开始循环并等待直到定时器触发退出循环
                if self.ifSynCtrlTriggerSuccess == 0:
                    self.CommunicationTextBrowser.append(f"周期采集中第{i+1}次采集中间板未能收到正确应答信号，建议检查")
                    self.SpectroscopyTextBrowser.append(f"周期采集中第{i+1}次采集中间板未能收到正确应答信号，建议检查")
                    break
                elif self.ifSynCtrlTriggerSuccess == 2:
                    self.retry_loop(self.AcquireData, 
                    "读数成功", 
                    "读数信号未接收到应答包，重新发送")
                    all_data.append(self.get_channel_data())
            
            # self.retry_loop(self.AcquireData, 
            #         "读数成功", 
            #         "读数信号未接收到应答包，重新发送")
            # all_data.append(self.get_channel_data())
            
            if self.PeriodIntervalEdit.text().strip() == "":
                PeriodInterval = 1000
            else:
                PeriodInterval = int(int(self.PeriodIntervalEdit.text())*1000)

            loop = QEventLoop()     # 创建一个循环
            QTimer.singleShot(PeriodInterval, loop.quit)  # 设置定时器，触发时退出循环
            loop.exec_()    # 开始循环并等待直到定时器触发退出循环
            
        self.FileSaveToExcel(all_data)        
                
    def get_channel_data(self): # 获取当前32个通道的数据并返回一个列表
        data = list(self.ChannelLongDATA.values())
        return data            

    def FileSaveToExcel(self, all_data):  # 文件保存到Excel
        FilePath = self.FilePathLineEdit.text()
        FileName = self.FileNameLineEdit.text()

        if not FileName.endswith('.xlsx'):
            FileName += '.xlsx'

        full_path = os.path.join(FilePath, FileName)

        # 将数据转换为 DataFrame
        df = pd.DataFrame(all_data).T

        try:
            # 检查并创建文件所在的目录
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
        except Exception as e:
            self.SpectroscopyTextBrowser.append("文件路径错误,无法创建：%s" % e)
            self.CommunicationTextBrowser.append("文件路径错误,无法创建：%s" % e)
            return

        # 打开或创建一个新的 Excel 文件
        new_file = False
        try:
            workbook = load_workbook(full_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            new_file = True

        # 写入每个周期的数据，并添加周期标签
        for col_index, col_data in enumerate(df.columns):
            sheet.cell(row=1, column=col_index + 1, value=f"周期 {col_index + 1}")
            for row_index, value in enumerate(df[col_data]):
                sheet.cell(row=row_index + 2, column=col_index + 1, value=value)

        # 保存 Excel 文件
        workbook.save(full_path)
        self.SpectroscopyTextBrowser.append("文件保存成功：%s" % full_path)
        self.CommunicationTextBrowser.append("文件保存成功：%s" % full_path)

    def SingleChannelThresholdTuning(self): #  单通道阈值微调
        ChannelNumberStr = self.ChannelNumberEdit.text()    # 获取通道号输入框中的文本，转换为整数
        ChannelNumberValue = int(ChannelNumberStr)          # 将通道号从字符串转换为整数
        
        ThresholdStr = self.ThresholdEdit.text()            # 获取阈值输入框中的文本，转换为整数
        ThresholdValue = int(ThresholdStr)                  # 将阈值从字符串转换为整数
        # 检查通道号是否在有效范围（1-64）
        if ChannelNumberValue > 64: 
            self.SpectroscopyTextBrowser.append("通道号超出范围")
            self.CommunicationTextBrowser.append("通道号超出范围")
            return
        
        # 检查阈值是否在有效范围（0-31）
        if ThresholdValue > 31: 
            self.SpectroscopyTextBrowser.append("阈值超出范围")
            self.CommunicationTextBrowser.append("阈值超出范围")
            return
        
        InsSingleChannelThresholdTuningCode = 0X02  # 定义指令码，并将其转换为两位十六进制字符串
        str_InsSingleChannelThresholdTuningCode = f"{InsSingleChannelThresholdTuningCode:02X}"
        
        SingleChannelThresholdTuning_combined = (0b11 << 14) | (ChannelNumberValue-1 << 8) | (0b000 << 5) | ThresholdValue    # 组合指令
        str_SingleChannelThresholdTuning_combined = f"{SingleChannelThresholdTuning_combined:04X}"  # 将组合后的指令转换为四位十六进制字符串
        str_SingleChannelThresholdTuning = str_InsSingleChannelThresholdTuningCode + str_SingleChannelThresholdTuning_combined  # 合并指令码和组合的指令数据，构成最终的指令字符串
        Instr_SingleChannelThresholdTuning = self.InstrCombination(str_SingleChannelThresholdTuning)
        self.send_data(Instr_SingleChannelThresholdTuning)
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWin = DataTransmission()
    myWin.show()
    sys.exit(app.exec_())
